#!/usr/bin/env python3
"""
Life Care Planner Directory - Data Import Script

Reads an Excel file (.xlsx) with Life Care Planner data and converts it
into a structured JSON file used by the Astro static site generator.

Usage:
    python3 scripts/import.py

Input:  LCP_Leads_With_City_State.xlsx (in project root)
Output: src/data/providers.json
"""

import json
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl

INPUT_FILE = "LCP_Leads_With_City_State.xlsx"
OUTPUT_FILE = Path("src") / "data" / "providers.json"

# State/province abbreviation to full slug name map
STATE_ABBREV_MAP = {
    # US States
    "AL": "alabama", "AK": "alaska", "AZ": "arizona", "AR": "arkansas",
    "CA": "california", "CO": "colorado", "CT": "connecticut", "DE": "delaware",
    "FL": "florida", "GA": "georgia", "HI": "hawaii", "ID": "idaho",
    "IL": "illinois", "IN": "indiana", "IA": "iowa", "KS": "kansas",
    "KY": "kentucky", "LA": "louisiana", "ME": "maine", "MD": "maryland",
    "MA": "massachusetts", "MI": "michigan", "MN": "minnesota", "MS": "mississippi",
    "MO": "missouri", "MT": "montana", "NE": "nebraska", "NV": "nevada",
    "NH": "new-hampshire", "NJ": "new-jersey", "NM": "new-mexico", "NY": "new-york",
    "NC": "north-carolina", "ND": "north-dakota", "OH": "ohio", "OK": "oklahoma",
    "OR": "oregon", "PA": "pennsylvania", "RI": "rhode-island",
    "SC": "south-carolina", "SD": "south-dakota", "TN": "tennessee",
    "TX": "texas", "UT": "utah", "VT": "vermont", "VA": "virginia",
    "WA": "washington", "WV": "west-virginia", "WI": "wisconsin", "WY": "wyoming",
    "DC": "district-of-columbia",
    # Canadian Provinces
    "AB": "alberta", "BC": "british-columbia", "MB": "manitoba",
    "NB": "new-brunswick", "NL": "newfoundland-and-labrador",
    "NS": "nova-scotia", "NT": "northwest-territories", "NU": "nunavut",
    "ON": "ontario", "PE": "prince-edward-island", "QC": "quebec",
    "SK": "saskatchewan", "YT": "yukon",
}

CANADIAN_PROVINCES = {
    "AB", "BC", "MB", "NB", "NL", "NS", "NT", "NU",
    "ON", "PE", "QC", "SK", "YT",
}


def slugify(text):
    """Convert any text to a URL-friendly lowercase slug."""
    if not text:
        return ""
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[-_\s]+", "-", text)
    return text.strip("-")


def normalize_phone(phone):
    """Clean and standardize phone numbers."""
    if not phone:
        return ""
    phone = str(phone).strip()
    # Remove leading country codes like "1-" or "1 "
    phone = re.sub(r"^1[-.\s]", "", phone)
    # Standardize dash spacing
    phone = re.sub(r"\s*-\s*", "-", phone)
    return phone


def parse_expiration_date(value):
    """Parse expiration date from various formats."""
    if not value:
        return ""
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        # Already ISO format or similar
        value = value.strip()
        if " " in value:
            value = value.split(" ")[0]  # Take date part only
        return value
    return str(value)[:10] if len(str(value)) >= 10 else str(value)


def resolve_country(state_code: str) -> str:
    """Determine country from state/province abbreviation."""
    if state_code in CANADIAN_PROVINCES:
        return "Canada"
    return "USA"


def resolve_state_name(state_code: str) -> str:
    """Convert state abbreviation to full display name."""
    slug = STATE_ABBREV_MAP.get(state_code.upper(), "")
    if slug:
        return slug  # Return slug form (e.g., 'new-york')
    return slugify(state_code) if state_code else ""


def parse_provider_row(row: dict, index: int) -> dict | None:
    """
    Parse a single row from the xlsx into a provider dictionary.
    Returns None if the row should be skipped (missing critical data).
    """
    name = (row.get("Name") or "").strip()
    company = (row.get("Company") or "").strip()
    address = (row.get("Address") or "").strip()
    zip_code = str(row.get("Zip") or "").strip()
    city = (row.get("City") or "").strip()
    state_code = (row.get("State") or "").strip().upper()
    phone = normalize_phone(row.get("Telephone") or "")
    email = (row.get("Email") or "").strip().lower()
    certification = (row.get("Type_of_Certification") or "").strip()
    expiration = parse_expiration_date(row.get("Expiration_Date"))
    cert_no = str(row.get("Certificate_No") or "").strip()
    url = (row.get("url") or "").strip()

    # Validate — skip rows without name
    if not name:
        return None

    # Resolve state slug and country
    state_slug = resolve_state_name(state_code)
    country = resolve_country(state_code) if state_code else "Unknown"

    # Handle missing state but has city
    if not state_slug and city:
        state_slug = "unknown"
        country = "Unknown"
    if state_slug and not city:
        city = "unknown"

    # Option A: Filter out rows with no state OR no city (after resolution)
    if not state_slug or state_slug == "unknown":
        return None
    if not city or city == "unknown":
        return None

    city_slug = slugify(city)

    # Build full address string
    addr_parts = [address] if address else []
    if city_slug:
        addr_parts.append(city)
    if state_code:
        addr_parts.append(state_code)
    if zip_code and zip_code != "None":
        addr_parts.append(zip_code)
    full_address = ", ".join(addr_parts) if addr_parts else ""

    # Generate description from certification
    description = certification if certification else f"Life Care Planner located in {city}, {state_code}"
    
    # Generate specialty field (use certification or default)
    specialty = certification if certification else "Life Care Planner"

    # Slug from name — make unique with index if needed
    provider_slug = slugify(name)
    if not provider_slug:
        provider_slug = f"provider-{index + 1}"

    return {
        "id": index + 1,
        "name": name,
        "company": company,
        "country": country,
        "state": state_slug,
        "stateCode": state_code,
        "city": city_slug,
        "cityDisplay": city,
        "address": full_address,
        "phone": phone,
        "email": email,
        "website": url,
        "certification": certification,
        "specialty": specialty,
        "expirationDate": expiration,
        "certificateNo": cert_no,
        "description": description,
        "image": f"https://ui-avatars.com/api/?name={slugify(name)}&background=1e3a8a&color=fff&size=200",
        "slug": provider_slug,
    }


def main():
    input_path = Path(INPUT_FILE)

    if not input_path.exists():
        print(f"Error: {input_path.name} not found in project root.")
        print(f"Place your Excel file at: {input_path.absolute()}")
        sys.exit(1)

    wb = openpyxl.load_workbook(input_path, read_only=True)
    ws = wb.active

    # Read headers from first row
    headers = [cell.value for cell in ws[1]]

    # Read all data rows
    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
    total_raw = len(raw_rows)

    # Parse each row
    providers = []
    skipped = 0
    for i, row_data in enumerate(raw_rows):
        row_dict = dict(zip(headers, row_data))
        parsed = parse_provider_row(row_dict, i)
        if parsed:
            providers.append(parsed)
        else:
            skipped += 1

    wb.close()

    if not providers:
        print("Error: No valid providers found after filtering.")
        sys.exit(1)

    # Deduplicate slugs
    seen_slugs = {}
    for p in providers:
        slug = p["slug"]
        if slug in seen_slugs:
            seen_slugs[slug] += 1
            p["slug"] = f"{slug}-{seen_slugs[slug]}"
        else:
            seen_slugs[slug] = 0

    # Summary
    countries = sorted(set(p["country"] for p in providers if p["country"]))
    states = sorted(set(p["state"] for p in providers if p["state"]))

    data = {
        "providers": providers,
        "metadata": {
            "total": len(providers),
            "totalRaw": total_raw,
            "skipped": skipped,
            "lastUpdated": date.today().isoformat(),
            "countries": countries,
            "states": states,
        },
    }

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f" Imported {len(providers)} providers from {input_path.name}")
    print(f"   Raw rows: {total_raw} | Skipped: {skipped} | Kept: {len(providers)}")
    print(f"   Countries: {', '.join(countries)}")
    print(f"   States/Provinces: {len(states)}")
    print(f"   Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
